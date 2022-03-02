'''
Starting Date: 2022.03.02
Ending Date: 2022.03.02
Coder: Chan Kai San
E-mail: u3556373@connect.hku.hk
Description: This script aims to filter out the day14 KO data and create the mutant files according to the filtered data of SaCas9 from Zero's NAR paper
Reference: -
'''

import openpyxl
#Read in the file#
workbook=openpyxl.load_workbook('zero.xlsx')
worksheet = workbook.worksheets[0]
rows = worksheet.max_row   #Obtain the amount of the data
column = worksheet.max_column #Obtain the amount of column
print(rows,column)



'''
#Filtering out the KO value on Day 14 with sgRNA ON1#
with open("7b_mutant.txt","w") as f:
    with open("7bscore.txt","w") as f1:
        for i in range (1,rows+1):
            day_check = worksheet.cell(row = i, column = 22).value
            #print(day_check)
            if day_check == "D14":
                sgRNA = worksheet.cell(row = i, column = 23).value
                if sgRNA == "ZRP7b":
                    mutant=[]
                    KO_1 = worksheet.cell(row = i, column = 24).value
                    for j in range (2,21):
                        mutant.append(worksheet.cell(row = i, column = j).value)
                    #print(len(mutant))
                    f.write("TA238%s,YA239%s,RA245%s,TA392%s,NA394%s,QA414%s,NA419%s,RA499%s,QA500%s,YA651%s,RA654%s,GA655%s,EA782%s,NA968%s,RA1015%s,NA260%s,NA413%s,GA616%s,NA888%s;\r\n"%(mutant[0],mutant[1],mutant[2],mutant[3],mutant[4],mutant[5],mutant[6],mutant[7],mutant[8],mutant[9],mutant[10],mutant[11],mutant[12],mutant[13],mutant[14],mutant[15],mutant[16],mutant[17],mutant[18]))
                    f1.write(str(KO_1)+'\r\n')
'''

#Filtering out the KO value on Day 14 with sgRNA ON2#
with open("7g_mutant.txt","w") as f:
    with open("7gscore.txt","w") as f1:
        for i in range (1,rows+1):
            day_check = worksheet.cell(row = i, column = 22).value
            #print(day_check)
            if day_check == "D14":
                sgRNA = worksheet.cell(row = i, column = 23).value
                if sgRNA == "ZRP7g":
                    mutant=[]
                    KO_1 = worksheet.cell(row = i, column = 24).value
                    for j in range (2,21):
                        mutant.append(worksheet.cell(row = i, column = j).value)
                    #print(len(mutant))
                    f.write("TA238%s,YA239%s,RA245%s,TA392%s,NA394%s,QA414%s,NA419%s,RA499%s,QA500%s,YA651%s,RA654%s,GA655%s,EA782%s,NA968%s,RA1015%s,NA260%s,NA413%s,GA616%s,NA888%s;\r\n"%(mutant[0],mutant[1],mutant[2],mutant[3],mutant[4],mutant[5],mutant[6],mutant[7],mutant[8],mutant[9],mutant[10],mutant[11],mutant[12],mutant[13],mutant[14],mutant[15],mutant[16],mutant[17],mutant[18]))
                    f1.write(str(KO_1)+'\r\n')


print("Your file is generating automatically......")



#T238	Y239	R245	T392	N394	Q414	N419	R499	Q500	Y651	R654	G655	E782	N968	R1015	N260	N413	G616	N888
