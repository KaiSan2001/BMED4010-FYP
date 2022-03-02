'''
Starting Date: 2022.02.25
Ending Date: 2022.02.25
Coder: Chan Kai San
E-mail: u3556373@connect.hku.hk
Description: This script aims to create the mutant files according to the filtered data of SpCas9

'''

import openpyxl
#Read in the file#
workbook=openpyxl.load_workbook('SpCas9.xlsx')

worksheet = workbook.worksheets[0]
rows = worksheet.max_row   #Obtain the amount of the data
print(rows)
print("Your file is generating automatically......")
AAList=['A','C','D','E','F','G','H','I','K','L','M','N','P','Q','R','S','T','V','W','Y']
with open("SpCas9_mutant.txt","w") as f:
    with open ("SpCas9_single.txt","w") as f1:
        for i in range(2,954):
            mutant = worksheet.cell(row=i,column=1).value
            fitness = str(worksheet.cell(row=i,column=2).value)
            if fitness != 'NA':
                for aa1 in AAList:
                    
                    desire = 'RQKET%sKR'%(aa1)
                
                    if mutant == desire:
                        print(mutant)
                        f.write("QB926%s;\r\n"%aa1)
                        f1.write(fitness+'\r\n')

'''
with open("SpCas9_mutant.txt","w") as f:
    with open ("SpCas9_single.txt","w") as f1:
        for i in range(2,954):
            mutant = worksheet.cell(row=i,column=1).value
            fitness = str(worksheet.cell(row=i,column=2).value)
            if fitness != 'NA':
                for aa1 in AAList:
                    for aa2 in AAList:
                        for aa3 in AAList:
                            desire = 'RQK%s%s%sKR'%(aa1,aa2,aa3)
                            if mutant == desire:
                                print(mutant)
                                f.write("EB923%s,TB924%s,QB926%s;\r\n"%(aa1,aa2,aa3))
                                f1.write(fitness+'\r\n')
                        #f.write("RB661%s,QB695%s,KB848%s,EB923%s,TB924%s,QB926%s,KB1003%s,RB1060%s;\r\n"%(mutant[0],mutant[1],mutant[2],mutant[3],mutant[4],mutant[5],mutant[6],mutant[7]))
path = os.path.abspath('.')
name = 'SpCas9_mutants.txt'
content='Your file "%s" had been saved into %s' %(name,path)
print('-'*len(content));print(content);print('-'*len(content))
print('\n')
'''
