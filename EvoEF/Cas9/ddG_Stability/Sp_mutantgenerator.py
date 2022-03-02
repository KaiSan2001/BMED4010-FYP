'''
Starting Date: 2022.02.19
Ending Date: 2022.02.19
Coder: Chan Kai San
E-mail: u3556373@connect.hku.hk
Description: This script aims to create the mutant files according to the filtered data of SaCas9

'''

import openpyxl
import os

#Read in the file#
workbook=openpyxl.load_workbook('1.xlsx')

worksheet = workbook.worksheets[0]
rows = worksheet.max_row   #Obtain the amount of the data
print(rows)
print("Your file is generating automatically......")
with open("SpCas9_mutant.txt","w") as f:
    for i in range(2,954):
        mutant = worksheet.cell(row=i,column=7).value
        f.write("RB661%s,QB695%s,KB848%s,EB923%s,TB924%s,QB926%s,KB1003%s,RB1060%s;\r\n"%(mutant[0],mutant[1],mutant[2],mutant[3],mutant[4],mutant[5],mutant[6],mutant[7]))
path = os.path.abspath('.')
name = 'SpCas9_mutants.txt'
content='Your file "%s" had been saved into %s' %(name,path)
print('-'*len(content));print(content);print('-'*len(content))
print('\n')