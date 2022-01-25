'''
Starting Date: 2022.01.24
Ending Date: 2022.01.24
Last Editing Date: 2022.01.25
Coder: Chan Kai San
E-mail: u3556373@connect.hku.hk
Description: This script aims to read in the filtered data from publication, divide the total amount into 10 groups, select 100 samples randomly from each group, and save the samples into
a new file for later operation
Reference: - Excel File processing: https://blog.csdn.net/sinat_28576553/article/details/81275650
           - Random Sampling: https://blog.csdn.net/qq_31279347/article/details/82795405
           - Dictionary: https://blog.csdn.net/a411178010/article/details/78548168
'''

import openpyxl
import random
import os

#Read in the file#
workbook=openpyxl.load_workbook('GB1.xlsx')

worksheet = workbook.worksheets[0]
#rows = worksheet.max_row   #Obtain the amount of the data, should be 149362

'''After sorting the whole data in ascending order, dividing the elements into 10 groups and randomly select 100 samples from each group. In total, we can obtain 1000 data points from each hierarchy.
Equivalently divide 149362 samples into 10 groups, so each of the group will approximately contain 14936 data. Randomly choose 100 points from it'''

#group_index = [14936, 29872, 44808, 59744, 74680, 89616, 104552, 119488, 134424, 149362]

index1 = random.sample(range(2,14936),100)#avoid to choose the head
index2 = random.sample(range(14936,29872),100)
index3 = random.sample(range(29872,44808),100)
index4 = random.sample(range(44808,59744),100)
index5 = random.sample(range(59744,74680),100)
index6 = random.sample(range(74680,89616),100)
index7 = random.sample(range(89616,104552),100)
index8 = random.sample(range(104552,119488),100)
index9 = random.sample(range(119488,134424),100)
index10 = random.sample(range(134424,149363),100) #include the last term
#index11 = random.sample(range(134424,149363),100) #will lead to the repeat mutation


data = {}

#operate the file#
index = [index1, index2, index3, index4, index5, index6, index7, index8, index9, index10]
for i in range(len(index)):
    for j in index[i]:
        mutant = worksheet.cell(row=j, column=1).value #Note that the operation only regard on the worksheet
        fitness = worksheet.cell(row=j, column=5).value
        data[mutant]=fitness



#Create a file to save data#
savebook = openpyxl.Workbook()
ws=savebook.active
ws.title='Filtered data'
ws.cell(1,1,'Mutant')
ws.cell(1,2,'Fitness')
i=2
while i < 1002:
    for key,value in data.items():
        ws.cell(i,1,key)
        ws.cell(i,2,value)
        i+=1
savebook.save('RandomData.xlsx')
path = os.path.abspath('.')
name = 'RandomData.xlsx'
content='Your file "%s" had been saved into %s' %(name,path)
print('-'*len(content));print(content);print('-'*len(content))
print('\n')
