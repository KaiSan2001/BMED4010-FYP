'''
Starting Date: 2022.01.25
Ending Date: 2022.01.25
Coder: Chan Kai San
E-mail: u3556373@connect.hku.hk
Description: This script aims to create the mutant files according to the filtered data
Reference: - Excel File processing 
           - Txt File processing
           - Mutant File formating
           - Mutant Site selection
'''

import openpyxl
import os

#Read in the file#
workbook=openpyxl.load_workbook('RandomData_5.xlsx')

worksheet = workbook.worksheets[0]
rows = worksheet.max_row   #Obtain the amount of the data
print(rows)
print("Your file is generating automatically......")
with open("2GI9_mutants_5.txt","w") as f:
    for i in range(2,rows+1):
        mutant = worksheet.cell(row=i,column=1).value
        f.write("VA39%s,DA40%s,GA41%s,VA54%s;\r\n"%(mutant[0],mutant[1],mutant[2],mutant[3]))
path = os.path.abspath('.')
name = '2GI9_mutants.txt'
content='Your file "%s" had been saved into %s' %(name,path)
print('-'*len(content));print(content);print('-'*len(content))
print('\n')